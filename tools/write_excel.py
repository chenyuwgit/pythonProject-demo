# -*- coding:utf-8 -*-
# 作者：NoamaNelson
# 日期：2021/8/7 9:15
# 文件名称：write_excel.py
# 作用：写入excel数据以及excel复制到指定目录
# 联系：VX(NoamaNelson)
# 博客：https://blog.csdn.net/NoamaNelson


from openpyxl import load_workbook
import openpyxl
from conf import config


def copy_excel(excelpath1, excelpath2):
    """复制excek，把excelpath1数据复制到excelpath2"""
    wb2 = openpyxl.Workbook()
    wb2.save(excelpath2)
    # 读取数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row                      # 最大行数
    max_column = sheet1.max_column                # 最大列数

    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):   # chr(97)='a'
            n = chr(n)                            # ASCII字符
            i ='%s%d'% (n, m)                     # 单元格编号
            cell1 = sheet1[i].value               # 获取data单元格数据
            sheet2[i].value = cell1               # 赋值到test单元格

    wb2.save(excelpath2)                          # 保存数据
    wb1.close()                                   # 关闭excel
    wb2.close()


class Write_excel(object):
    """修改excel数据"""
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        """写入数据，如(2,3，"python"),第二行第三列写入数据"python"""
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)


if __name__ == "__main__":
    aa = Write_excel(config.CASE_EXCEL_PATH)
    aa.write(6, 6, "人生苦短，我用python！")
